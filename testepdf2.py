import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os

def extrair_dados_pdf(pdf_path):
    dados_extraidos = {
        "CCM": "", "CNPJ": "", "Mês Referência": "", "Situação": "", "Encerramento": "",
        "Endereço": "", "Número": "", "Complemento": "", "Bairro": "", "Cidade": "", "Estado": "",
        "Lançamentos Válidos": [], "Lançamentos Substituídos": [], "Encerramentos": []
    }
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            lines = text.split("\n")
            
            is_valid = None  # Inicializa a variável is_valid
            
            for line in lines:
                if "CCM:" in line:
                    parts = line.split()
                    dados_extraidos["CCM"] = parts[1]
                    dados_extraidos["CNPJ"] = parts[3]
                    dados_extraidos["Mês Referência"] = parts[5]
                    dados_extraidos["Situação"] = parts[7]
                    if len(parts) > 9:  # Verifica se há elementos suficientes
                        dados_extraidos["Encerramento"] = parts[9]
                    else:
                        dados_extraidos["Encerramento"] = ""  # Ou algum valor padrão
                
                elif "Endereço:" in line:
                    parts = line.split(":")
                    dados_extraidos["Endereço"] = parts[1].strip()
                elif "Número:" in line:
                    dados_extraidos["Número"] = line.split(":")[1].strip()
                elif "Complemento:" in line:
                    dados_extraidos["Complemento"] = line.split(":")[1].strip()
                elif "Bairro:" in line:
                    dados_extraidos["Bairro"] = line.split(":")[1].strip()
                elif "Cidade:" in line:
                    dados_extraidos["Cidade"] = line.split(":")[1].strip()
                elif "Estado:" in line:
                    dados_extraidos["Estado"] = line.split(":")[1].strip()
                
                elif "LANÇAMENTOS VÁLIDOS" in line:
                    is_valid = True
                elif "LANÇAMENTOS SUBSTITUÍDOS" in line:
                    is_valid = False
                elif "HISTORICO DE ENCERRAMENTOS" in line:
                    is_valid = None
                
                elif is_valid is True and len(line.split()) > 5:
                    dados_extraidos["Lançamentos Válidos"].append(line.split())
                elif is_valid is False and len(line.split()) > 5:
                    dados_extraidos["Lançamentos Substituídos"].append(line.split())
                elif is_valid is None and len(line.split()) > 2:
                    dados_extraidos["Encerramentos"].append(line.split())
    
    return dados_extraidos

def criar_planilha(dados, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livro Fiscal"
    
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thick_border = Border(left=Side(style="thick"), right=Side(style="thick"), top=Side(style="thick"), bottom=Side(style="thick"))
    
    # Cabeçalho
    ws.append(["Prefeitura Municipal de Itapira"])
    ws.append(["SECRETARIA MUNICIPAL DA FAZENDA"])
    ws.append(["Divisão de ISSQN"])
    ws.append([""])
    ws.append(["Livro Fiscal Serviços Prestados Mensal"])
    ws.append(["Data impressão"])
    ws.append([""])
    
    ws.append(["CCM:", dados["CCM"], "CPF / CNPJ:", dados["CNPJ"], "Mês Referência:", dados["Mês Referência"], "Situação:", dados["Situação"], "Encerramento:", dados["Encerramento"]])
    ws.append([""])
    ws.append(["Endereço:", dados["Endereço"], "Número:", dados["Número"], "Complemento:", dados["Complemento"], "Bairro:", dados["Bairro"], "Cidade:", dados["Cidade"], "Estado:", dados["Estado"]])
    ws.append([""])
    
    ws.append(["LANÇAMENTOS VÁLIDOS"])
    ws.append(["Dia", "Número", "RPS", "Série", "Tipo", "Situação", "Cod.", "Aliq.(%)", "Valor(R$)", "Base(R$)", "ISS", "ISS Retido", "ISS Trib. Fora", "ISS Retido Fora", "CNPJ Tomador", "Razão Tomador", "Lanç.", "Data Escrituração"])
    
    for row in dados["Lançamentos Válidos"]:
        ws.append(row)
    ws.append(["Total"])
    
    ws.append(["LANÇAMENTOS SUBSTITUÍDOS"])
    ws.append(["Dia", "Número", "RPS", "Série", "Tipo", "Situação", "Cod.", "Aliq.(%)", "Valor(R$)", "Base(R$)", "ISS", "ISS Retido", "ISS Trib. Fora", "ISS Retido Fora", "CNPJ Tomador", "Razão Tomador", "Evento", "Data Escrituração"])
    
    for row in dados["Lançamentos Substituídos"]:
        ws.append(row)
    
    ws.append(["HISTORICO DE ENCERRAMENTOS DO LIVRO"])
    for row in dados["Encerramentos"]:
        ws.append(row)
    
    # Formatação
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
            if cell.row <= 12 or cell.row == 15 or cell.row == len(ws["A"]):
                cell.font = bold_font
            if cell.value:
                cell.border = thick_border
    
    wb.save(output_path)
    print(f"Arquivo Excel salvo em: {output_path}")

def main(pdf_path, output_path):
    dados = extrair_dados_pdf(pdf_path)
    criar_planilha(dados, output_path)

if __name__ == "__main__":
    pdf_file = r"C:\Users\bruno.martins\Desktop\ResolveCaptcha\livro fiscal\teste.pdf"
    excel_file = r"C:\Users\bruno.martins\Desktop\ResolveCaptcha\livro fiscal\resultado.xlsx"
    main(pdf_file, excel_file)
