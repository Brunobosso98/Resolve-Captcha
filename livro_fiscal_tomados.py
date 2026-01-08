import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException
import pytesseract
from PIL import Image
import pandas as pd
import time
import os
import glob
import re
import json
import shutil
from openpyxl import load_workbook


def get_resource_path(relative_path):
    """Obtém o caminho absoluto para recursos, funciona para desenvolvimento e para PyInstaller."""
    if hasattr(sys, '_MEIPASS'):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def click_element(wait, locator, descricao, tentativas=3):
    """Clica em um elemento garantindo nova referência quando ocorrer StaleElement."""
    ultima_excecao = None
    for _ in range(tentativas):
        try:
            elemento = wait.until(EC.element_to_be_clickable(locator))
            elemento.click()
            print(f"{descricao} clicado com sucesso!")
            return elemento
        except StaleElementReferenceException as exc:
            ultima_excecao = exc
            time.sleep(0.4)
    raise ultima_excecao or Exception("Não foi possível clicar no elemento.")


def construir_pasta_livro(ano, mes):
    """Retorna o caminho estruturado dentro de Livro e garante que exista."""
    pasta = os.path.join(os.getcwd(), "Livro", str(ano), str(mes).zfill(2))
    os.makedirs(pasta, exist_ok=True)
    return pasta


# Configuração idêntica à do captcha2.py para evitar divergências
CAMINHO_TESSERACT = r'W:\Fiscal\Escrita Fiscal\Davi\dependencias sistema\Tesseract-OCR\tesseract.exe'
CAMINHO_EXCEL = get_resource_path('Senha Municipio Itapira.xlsx')
URL_LOGIN = 'https://itapira.sigiss.com.br/itapira/contribuinte/login.php'

pytesseract.pytesseract.tesseract_cmd = CAMINHO_TESSERACT


def extrair_numeros_imagem(driver, wait):
    numeros = None
    try:
        elemento_imagem = wait.until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[1]/div[2]/div/div[2]/div[4]/div/div/div/span/img'))
        )

        driver.save_screenshot('screenshot.png')
        screenshot = Image.open('screenshot.png')

        location = elemento_imagem.location
        size = elemento_imagem.size
        left = location['x']
        top = location['y']
        right = location['x'] + size['width']
        bottom = location['y'] + size['height']

        imagem = screenshot.crop((left, top, right, bottom))
        imagem.save('numero.png')

        imagem = imagem.convert('L')
        numeros = pytesseract.image_to_string(imagem, config='--psm 6 -c tessedit_char_whitelist=0123456789')
        numeros = ''.join(filter(str.isdigit, numeros))

    except Exception as e:
        print(f"Erro na extração: {type(e).__name__}: {e}")

    return numeros


def preencher_campo(driver, element_id, valor, wait):
    try:
        campo = wait.until(
            EC.element_to_be_clickable((By.ID, element_id))
        )
        campo.clear()
        campo.send_keys(valor)
        print(f"Campo {element_id} preenchido")
    except Exception as e:
        print(f"Erro ao preencher {element_id}: {e}")


def processar_login(driver, wait):
    try:
        try:
            WebDriverWait(driver, 8).until(lambda d: d.current_url != URL_LOGIN)
        except Exception:
            pass

        current_url = driver.current_url
        if 'main.php' in current_url:
            return True

        if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
            print("Erro de login: Código de Confirmação Inválido (captcha)")
            return False

        if "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
            print("Erro de login: Contribuinte Inexistente ou Senha Inválida")
            return False

        try:
            erro_elemento = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[2]/div/font/b/center')))
            if erro_elemento.is_displayed():
                print("Erro no login: ", erro_elemento.text)
                if "Confirmação" in erro_elemento.text or "confirmação" in erro_elemento.text or "Código" in erro_elemento.text:
                    return False
                return False
        except Exception:
            pass

        if "login.php" not in current_url:
            return True

    except Exception as e:
        print(f"Erro no processo de login: {e}")
        return False

    return False


def digitar_captcha(driver, numeros, wait):
    try:
        campo = wait.until(
            EC.element_to_be_clickable((By.ID, "confirma"))
        )
        campo.clear()
        campo.send_keys(numeros)
        print("Captcha digitado com sucesso!")

        botao_logar = wait.until(EC.element_to_be_clickable((By.ID, "btnOk")))
        botao_logar.click()

    except Exception as e:
        print(f"Erro ao digitar captcha: {e}")


def clicar_livro_fiscal(driver, wait, mes, ano, empresa):
    try:
        click_element(wait, (By.XPATH, "//button[contains(text(), 'Serviços Tomados')]"), "Botão 'Serviços Tomados'")
        click_element(wait, (By.XPATH, "//a[contains(@onclick, 'livroMensalT')]"), "Link 'Livro Fiscal'")

        time.sleep(3)
        janelas = driver.window_handles
        if len(janelas) > 1:
            driver.switch_to.window(janelas[1])
            print("Mudou para a nova aba com o PDF do Livro Fiscal.")

            pasta_download = construir_pasta_livro(ano, mes)

            time.sleep(5)

            arquivos_pdf = glob.glob(os.path.join(pasta_download, "*.pdf"))
            if arquivos_pdf:
                arquivos_pdf.sort(key=os.path.getmtime, reverse=True)
                arquivo_original = arquivos_pdf[0]

                nome_limpo = re.sub(r'[<>:"/\\|?*]', '_', empresa)
                novo_nome = os.path.join(pasta_download, f"{nome_limpo}.pdf")

                shutil.move(arquivo_original, novo_nome)
                print(f"Arquivo PDF movido para: {novo_nome}")
            else:
                print("Nenhum arquivo PDF encontrado na pasta de download.")

            driver.close()
            print("Fechou a aba do PDF do Livro Fiscal.")

            driver.switch_to.window(janelas[0])
            print("Retornou para a aba principal.")
        else:
            print("Não foi possível abrir o PDF do Livro Fiscal em uma nova aba.")

    except Exception as e:
        print(f"Botão 'Serviços Tomados' não encontrado ou erro ao clicar em 'Livro Fiscal': {e}")


def preencher_data(driver, wait, mes, ano, empresa):
    try:
        campo_modificar = wait.until(
            EC.element_to_be_clickable((By.ID, "btnAlterar"))
        )
        campo_modificar.click()

        campo_mes = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[3]/select')))
        campo_mes.send_keys(mes)
        print(f"Mês '{mes}' digitado com sucesso!")

        campo_ano = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[7]/input')))
        campo_ano.clear()
        campo_ano.send_keys(ano)
        print(f"Ano '{ano}' digitado com sucesso!")

        botao_ok = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "btn-success")))
        botao_ok.click()
        print("Botão OK clicado com sucesso!")
        time.sleep(1)

        driver.refresh()
        time.sleep(3)

        clicar_livro_fiscal(driver, wait, mes, ano, empresa)

    except Exception as e:
        print(f"Erro ao preencher data: {e}")


def atualizar_excel_status(linha_index, mensagem):
    """Atualiza a coluna 'Status Processo' no Excel."""
    try:
        workbook = load_workbook(CAMINHO_EXCEL)
        worksheet = workbook.active

        coluna_status = None
        for col_idx, col_name in enumerate(worksheet[1], 1):
            if col_name.value == 'Status Processo':
                coluna_status = col_idx
                break

        if coluna_status is not None:
            worksheet.cell(row=linha_index + 2, column=coluna_status, value=mensagem)
        else:
            coluna_status = worksheet.max_column + 1
            worksheet.cell(row=1, column=coluna_status, value='Status Processo')
            worksheet.cell(row=linha_index + 2, column=coluna_status, value=mensagem)

        workbook.save(CAMINHO_EXCEL)
        workbook.close()
    except Exception as e:
        print(f"Erro ao atualizar o Excel: {e}")


def main():
    try:
        df = pd.read_excel(CAMINHO_EXCEL, engine='openpyxl')

        for index, row in df.iterrows():
            tentativas = 0
            max_tentativas = 8
            login_bem_sucedido = False
            login_falhou_credenciais = False

            while tentativas < max_tentativas and not login_bem_sucedido:
                chrome_options = Options()
                chrome_options.add_argument("--kiosk-printing")

                pasta_download = construir_pasta_livro(row['Ano'], row['Mês'])

                app_state = {
                    "recentDestinations": [
                        {"id": "Save as PDF", "origin": "local", "account": ""}
                    ],
                    "selectedDestinationId": "Save as PDF",
                    "version": 2
                }

                prefs = {
                    "download.default_directory": pasta_download,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "plugins.plugins_disabled": ["Chrome PDF Viewer"],
                    "plugins.plugin_field_trial_triggered": False
                }
                prefs["printing.print_preview_sticky_settings.appState"] = json.dumps(app_state)
                chrome_options.add_experimental_option("prefs", prefs)

                driver = webdriver.Chrome(options=chrome_options)
                driver.maximize_window()
                wait = WebDriverWait(driver, 20)
                driver.get(URL_LOGIN)

                time.sleep(2)

                try:
                    btn_ciente = wait.until(
                        EC.element_to_be_clickable((By.ID, "btnCiente"))
                    )
                    btn_ciente.click()
                    print("Botão 'Estou Ciente' clicado com sucesso!")
                except Exception:
                    print("Botão 'Estou Ciente' não encontrado ou já foi fechado.")

                try:
                    if tentativas == 0:
                        print(f"Processando linha {index + 1}: {row['Empresa']}")
                    else:
                        print(f"Tentativa {tentativas + 1} para a empresa: {row['Empresa']}")

                    preencher_campo(driver, "cnpj", row['Usuário'], wait)
                    preencher_campo(driver, "senha", row['Senha'], wait)

                    numeros = extrair_numeros_imagem(driver, wait)

                    if numeros:
                        print(f"Números extraídos: {numeros}")
                        digitar_captcha(driver, numeros, wait)

                        if processar_login(driver, wait):
                            login_bem_sucedido = True
                            preencher_data(driver, wait, row['Mês'], row['Ano'], row['Empresa'])
                        else:
                            current_url = driver.current_url
                            if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
                                print(f"Captcha inválido para {row['Empresa']}, tentando novamente...")
                                tentativas += 1
                            elif "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
                                print(f"Login falhou por credenciais incorretas para {row['Empresa']}")
                                atualizar_excel_status(index, 'Não foi possivel realizar o login.')
                                login_falhou_credenciais = True
                                break
                            else:
                                print("Login falhou por outro motivo, não preenchendo a data.")
                                break
                    else:
                        print("Nenhum número foi detectado!")
                        tentativas += 1
                        if tentativas < max_tentativas:
                            print(f"Tentando novamente ({tentativas}/{max_tentativas})...")
                finally:
                    driver.quit()

            if not login_bem_sucedido and not login_falhou_credenciais:
                print(f"Excedido o número máximo de tentativas para {row['Empresa']}. Indo para a próxima empresa.")

    except Exception as e:
        print(f"Erro geral: {e}")


if __name__ == "__main__":
    main()
