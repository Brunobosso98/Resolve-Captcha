import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import pytesseract
from PIL import Image
import pandas as pd
import time
import os
from openpyxl import load_workbook


def get_resource_path(relative_path):
    """Retorna caminho absoluto útil para PyInstaller e scripts normais."""
    if hasattr(sys, '_MEIPASS'):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# Configuração idêntica à do captcha2.py para evitar divergências
CAMINHO_TESSERACT = r'W:\Fiscal\Escrita Fiscal\Davi\dependencias sistema\Tesseract-OCR\tesseract.exe'
CAMINHO_EXCEL = get_resource_path('Senha Municipio Itapira.xlsx')
URL_LOGIN = 'https://itapira.sigiss.com.br/itapira/contribuinte/login.php'

pytesseract.pytesseract.tesseract_cmd = CAMINHO_TESSERACT


def extrair_numeros_imagem(driver, wait):
    """Captura visualmente o captcha e aplica OCR de dígitos."""
    numeros = None
    try:
        elemento_imagem = wait.until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[1]/div[2]/div/div[2]/div[4]/div/div/div/span/img'))
        )
        driver.save_screenshot('screenshot.png')
        screenshot = Image.open('screenshot.png')
        location = elemento_imagem.location
        size = elemento_imagem.size
        left = location['x']
        top = location['y']
        right = left + size['width']
        bottom = top + size['height']
        imagem = screenshot.crop((left, top, right, bottom))
        imagem = imagem.convert('L')
        numeros = pytesseract.image_to_string(imagem, config='--psm 6 -c tessedit_char_whitelist=0123456789')
        numeros = ''.join(filter(str.isdigit, numeros))
    except Exception as e:
        print(f"Erro na extração: {type(e).__name__}: {e}")
    return numeros


def preencher_campo(driver, element_id, valor, wait):
    """Preenche um campo identificado pelo ID."""
    try:
        campo = wait.until(EC.element_to_be_clickable((By.ID, element_id)))
        campo.clear()
        campo.send_keys(valor)
        print(f"Campo {element_id} preenchido")
    except Exception as e:
        print(f"Erro ao preencher {element_id}: {e}")


def processar_login(driver, wait):
    """Confirma se o login foi bem-sucedido e identifica erros clássicos."""
    try:
        if driver.current_url == 'https://itapira.sigiss.com.br/itapira/contribuinte/main.php':
            return True

        current_url = driver.current_url
        if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
            print("Erro de login: Código de Confirmação Inválido (captcha)")
            return False
        if "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
            print("Erro de login: Contribuinte Inexistente ou Senha Inválida")
            return False

        try:
            erro_elemento = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[2]/div/font/b/center')))
            if erro_elemento.is_displayed():
                print("Erro no login: ", erro_elemento.text)
                return False
        except Exception:
            pass
    except Exception as e:
        print(f"Erro no processo de login: {e}")
    return False


def digitar_captcha(driver, numeros, wait):
    try:
        campo = wait.until(EC.element_to_be_clickable((By.ID, "confirma")))
        campo.clear()
        campo.send_keys(numeros)
        botao_logar = wait.until(EC.element_to_be_clickable((By.ID, "btnOk")))
        botao_logar.click()
        print("Captcha digitado com sucesso!")
    except Exception as e:
        print(f"Erro ao digitar captcha: {e}")


def clicar_encerramento_fiscal_basico(driver, wait, mes, ano, empresa, linha_index=None):
    """Executa o Encerramento Fiscal até o alerta e retorna para o frame principal."""
    try:
        servicos_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Serviços Prestados')]")))
        servicos_btn.click()
        print("Botão 'Serviços Prestados' clicado com sucesso!")

        encerramento_link = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@onclick, 'fechamento/prestado.php')]")))
        encerramento_link.click()
        print("Link 'Encerramento Fiscal' clicado com sucesso!")

        time.sleep(4)
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "main")))
        print("Entrou no iframe 'main'.")

        botao_encerrar = wait.until(EC.element_to_be_clickable((By.ID, "btnSalvar")))
        botao_encerrar.click()
        print("Botão 'Encerrar Mes' clicado com sucesso!")

        try:
            fallback_wait = WebDriverWait(driver, 120)
            alert = fallback_wait.until(EC.alert_is_present())
            alert_text = alert.text
            print(f"Alerta encontrado: {alert_text}")
            alert.accept()
            print("Alerta aceito com sucesso!")
        except Exception:
            print("Nenhum alerta encontrado ou não foi possível interagir com ele após 2 minutos.")
            if linha_index is not None:
                atualizar_excel_status(linha_index, 'Escrituração pode não ter sido finalizada')
        driver.switch_to.default_content()
        print("Retornou ao frame principal.")
    except Exception as exc:
        print(f"Erro ao realizar encerramento fiscal: {exc}")


def preencher_data(driver, wait, mes, ano, empresa, linha_index=None):
    """Seleciona mês e ano e dispara o encerramento fiscal básico."""
    try:
        btn_alterar = wait.until(EC.element_to_be_clickable((By.ID, "btnAlterar")))
        btn_alterar.click()

        campo_mes = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[3]/select')))
        campo_mes.send_keys(mes)
        print(f"Mês '{mes}' digitado com sucesso!")

        campo_ano = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[7]/input')))
        campo_ano.clear()
        campo_ano.send_keys(ano)
        print(f"Ano '{ano}' digitado com sucesso!")

        botao_ok = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "btn-success")))
        botao_ok.click()
        print("Botão OK clicado com sucesso!")

        clicar_encerramento_fiscal_basico(driver, wait, mes, ano, empresa, linha_index)
    except Exception as e:
        print(f"Erro ao preencher data: {e}")


def atualizar_excel_status(linha_index, mensagem):
    """Atualiza o Excel com o status fornecido."""
    try:
        workbook = load_workbook(CAMINHO_EXCEL)
        worksheet = workbook.active
        coluna_status = None
        for col_idx, col_name in enumerate(worksheet[1], 1):
            if col_name.value == 'Status Processo':
                coluna_status = col_idx
                break
        if coluna_status is None:
            coluna_status = worksheet.max_column + 1
            worksheet.cell(row=1, column=coluna_status, value='Status Processo')
        worksheet.cell(row=linha_index + 2, column=coluna_status, value=mensagem)
        workbook.save(CAMINHO_EXCEL)
        workbook.close()
    except Exception as exc:
        print(f"Erro ao atualizar o Excel: {exc}")


def main():
    try:
        df = pd.read_excel(CAMINHO_EXCEL, engine='openpyxl')
        for index, row in df.iterrows():
            tentativas = 0
            max_tentativas = 8
            login_bem_sucedido = False
            login_falhou_credenciais = False

            while tentativas < max_tentativas and not login_bem_sucedido:
                chrome_options = Options()
                pasta_download = os.path.join(os.getcwd(), str(row['Ano']), str(row['Mês']).zfill(2))
                os.makedirs(pasta_download, exist_ok=True)
                prefs = {
                    "download.default_directory": pasta_download,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "plugins.plugins_disabled": ["Chrome PDF Viewer"],
                    "plugins.plugin_field_trial_triggered": False
                }
                chrome_options.add_experimental_option("prefs", prefs)
                driver = webdriver.Chrome(options=chrome_options)
                driver.maximize_window()
                wait = WebDriverWait(driver, 20)
                driver.get(URL_LOGIN)
                time.sleep(2)
                try:
                    btn_ciente = wait.until(EC.element_to_be_clickable((By.ID, "btnCiente")))
                    btn_ciente.click()
                    print("Botão 'Estou Ciente' clicado com sucesso!")
                except Exception:
                    print("Botão 'Estou Ciente' não encontrado ou já foi fechado.")

                try:
                    if tentativas == 0:
                        print(f"Processando linha {index + 1}: {row['Empresa']}")
                    else:
                        print(f"Tentativa {tentativas + 1} para a empresa: {row['Empresa']}")

                    preencher_campo(driver, "cnpj", row['Usuário'], wait)
                    preencher_campo(driver, "senha", row['Senha'], wait)

                    numeros = extrair_numeros_imagem(driver, wait)
                    if numeros:
                        print(f"Números extraídos: {numeros}")
                        digitar_captcha(driver, numeros, wait)
                        time.sleep(10)
                        if processar_login(driver, wait):
                            login_bem_sucedido = True
                            preencher_data(driver, wait, row['Mês'], row['Ano'], row['Empresa'], index)
                        else:
                            current_url = driver.current_url
                            if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
                                print(f"Captcha inválido para {row['Empresa']}, tentando novamente...")
                                tentativas += 1
                            elif "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
                                print(f"Login falhou por credenciais incorretas para {row['Empresa']}")
                                atualizar_excel_status(index, 'Não foi possivel realizar o login.')
                                login_falhou_credenciais = True
                                break
                            else:
                                print("Login falhou por outro motivo, não preenchendo a data.")
                                break
                    else:
                        print("Nenhum número foi detectado!")
                        tentativas += 1
                        if tentativas < max_tentativas:
                            print(f"Tentando novamente ({tentativas}/{max_tentativas})...")
                finally:
                    driver.quit()

            if not login_bem_sucedido and not login_falhou_credenciais:
                print(f"Excedido o número máximo de tentativas para {row['Empresa']}. Indo para a próxima empresa.")
    except Exception as exc:
        print(f"Erro geral: {exc}")


if __name__ == "__main__":
    main()
