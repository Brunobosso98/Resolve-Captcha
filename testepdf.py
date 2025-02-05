import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Caminho do arquivo formatado
file_path = "/mnt/data/livro_mensal_prestado_formatado.xlsx"

# Carregar a planilha
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Definir alinhamento centralizado e quebra de texto automática
alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Criar bordas espessas para todas as células preenchidas
border_style = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick")
)

# Iterar sobre todas as células preenchidas para aplicar formatação
for row in ws.iter_rows():
    for cell in row:
        if cell.value:
            cell.alignment = alignment
            cell.border = border_style

# Autoajustar largura das colunas
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)  # Obter a letra da coluna
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Salvar o arquivo com as novas formatações
formatted_file_path = "/mnt/data/livro_mensal_prestado_formatado_final.xlsx"
wb.save(formatted_file_path)

# Retornar o caminho do arquivo formatado
formatted_file_path

# Preciso transformar um arquivo PDF em Excel, e manipular seu layout é possíveo?.
# Todas as colunas que tiverem algum valor, devem: alinhar embaixo, centralizar, quebrar texto automaticamente, mesclar e centralizar e ter bordas externas espessas.
# Serão extraidos valores do PDF e os itens que terem esses valores, devem estar em NEGRITO no Excel: (CCM, CPF / CNPJ, Mês Referência, Situação, Encerramento, Razão Social:, Endereço:, Número:, Complemento:, Bairro:, Cidade:, Estado:, LANÇAMENTOS VÁLIDOS, Dia, Número, RPS, Série, Tipo, Situação, Cod., Aliq.(%), Valor(R$), Base(R$), ISS, ISS Retido, ISS Trib. Fora, ISS Retido Fora, CNPJ Tomador, Razão Tomador, Lanç, Data Escrituração, Total, LANÇAMENTOS SUBSTITUÍDOS, HISTORICO DE ENCERRAMENTOS DO LIVRO