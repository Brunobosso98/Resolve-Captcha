import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment

def extrair_texto_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        texto = []
        for page in pdf.pages:
            texto.append(page.extract_text())
    return texto

def criar_planilha(texto, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for linha in texto:
        ws.append([linha])  # Adiciona a linha diretamente na planilha

    # Modificando o layout
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obter a letra da coluna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width  # Ajusta a largura da coluna

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    wb.save(output_path)
    print(f"Arquivo Excel salvo em: {output_path}")

def ler_planilha_excel(caminho_arquivo):
    # Carrega o arquivo Excel
    wb = openpyxl.load_workbook(caminho_arquivo)
    # Seleciona a primeira planilha
    ws = wb.active
    
    # Lê os dados da planilha
    dados = []
    for linha in ws.iter_rows(values_only=True):
        dados.append(linha)
    
    return dados

def main(pdf_path, output_path):
    texto = extrair_texto_pdf(pdf_path)
    criar_planilha(texto, output_path)

if __name__ == "__main__":
    pdf_file = r"C:\Users\bruno.martins\Desktop\ResolveCaptcha\livro fiscal\teste.pdf"
    excel_file = r"C:\Users\bruno.martins\Desktop\ResolveCaptcha\livro fiscal\resultado.xlsx"
    main(pdf_file, excel_file)
