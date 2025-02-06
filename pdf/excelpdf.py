import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def extract_text_from_pdf(pdf_path, output_excel):
    with pdfplumber.open(pdf_path) as pdf:
        extracted_text = []
        for page in pdf.pages:
            extracted_text.extend(page.extract_text().split("\n"))
    
    # Criar um novo workbook
    wb = Workbook()
    ws = wb.active

    # Mesclar e centralizar as três primeiras linhas
    ws.merge_cells("A1:R1")
    ws["A1"] = extracted_text[0]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:R2")
    ws["A2"] = extracted_text[1]
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:R3")
    ws["A3"] = extracted_text[2]
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # Criar os cabeçalhos na linha 10
    headers = {
        "A10:D10": "CCM:",
        "E10:H10": "CPF / CNPJ:",
        "I10:K10": "Mês Referência:",
        "L10:N10": "Situação:",
        "O10:R10": "Encerramento:",
        "A12:R12": "Razão Social:",
        "A14:I14": "Endereço:",
        "J14:R14": "Número:",
        "A16:D16": "Complemento:",
        "E16:H16": "Bairro:",
        "I16:M16": "Cidade:",
        "N16:R16": "Estado:",
    }

    for cell_range, text in headers.items():
        ws.merge_cells(cell_range)
        ws[cell_range.split(":")[0]] = text
        ws[cell_range.split(":")[0]].alignment = Alignment(horizontal="center", vertical="center")

    # Ajuste para evitar erro de unpacking
    values = extracted_text[4].split(" ")
    ccm = values[0] if len(values) > 0 else ""
    cnpj = values[1] if len(values) > 1 else ""
    mes_referencia = values[2] if len(values) > 2 else ""
    situacao = " ".join(values[3:5]) if len(values) > 4 else ""
    encerramento = " ".join(values[5:]) if len(values) > 5 else ""

    # Encontrar os campos de Razão Social, Endereço, Número, Complemento, Bairro, Cidade, Estado
    def find_value(label):
        index = extracted_text.index(label) + 1 if label in extracted_text else -1
        return extracted_text[index] if index > 0 else ""

    razao_social = find_value("Razão Social:")
    endereco = find_value("Endereço:")
    numero = find_value("Número:")
    complemento = find_value("Complemento:")
    bairro = find_value("Bairro:")
    cidade = find_value("Cidade:")
    estado = find_value("Estado:")

    # Inserir os valores abaixo dos cabeçalhos
    values_map = {
        "A11:D11": ccm,
        "E11:H11": cnpj,
        "I11:K11": mes_referencia,
        "L11:N11": situacao,
        "O11:R11": encerramento,
        "A13:R13": razao_social,
        "A15:I15": endereco,
        "J15:R15": numero,
        "A17:D17": complemento,
        "E17:H17": bairro,
        "I17:M17": cidade,
        "N17:R17": estado,
    }

    for cell_range, value in values_map.items():
        ws.merge_cells(cell_range)
        ws[cell_range.split(":")[0]] = value
        ws[cell_range.split(":")[0]].alignment = Alignment(horizontal="center", vertical="center")

    # Inserir o restante do texto extraído no Excel a partir da linha 18
    row_num = 18
    for line in extracted_text[6:]:  # Alterado para evitar sobrescrita da linha 5
        ws.cell(row=row_num, column=1, value=line)
        row_num += 1

    # Salvar o arquivo Excel formatado
    wb.save(output_excel)
    print(f'Arquivo salvo como {output_excel}')

# Exemplo de uso
pdf_path = r"C:\Users\bruno.martins\Desktop\Automações\ResolveCaptcha\livro fiscal\teste.pdf"  # Substitua pelo caminho do seu PDF
output_excel = r"C:\Users\bruno.martins\Desktop\Automações\ResolveCaptcha\livro fiscal\resultado.xlsx"   # Nome do arquivo Excel de saída
extract_text_from_pdf(pdf_path, output_excel)
