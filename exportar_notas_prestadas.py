import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException
import pytesseract
from PIL import Image
import pandas as pd
import time
import os
import glob
import re
import zipfile
from openpyxl import load_workbook


def get_resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def click_element(wait, locator, descricao, tentativas=3):
    ultima_excecao = None
    for tentativa in range(tentativas):
        try:
            elemento = wait.until(EC.element_to_be_clickable(locator))
            elemento.click()
            print(f"{descricao} clicado com sucesso!")
            return elemento
        except StaleElementReferenceException as exc:
            ultima_excecao = exc
            time.sleep(0.4)
    raise ultima_excecao or Exception("Não foi possível clicar no elemento.")


CAMINHO_TESSERACT = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
CAMINHO_EXCEL = get_resource_path('Senha Municipio Itapira Prestadoras (Maria).xlsx')
URL_LOGIN = 'https://itapira.sigiss.com.br/itapira/contribuinte/login.php'

pytesseract.pytesseract.tesseract_cmd = CAMINHO_TESSERACT


def extrair_numeros_imagem(driver, wait):
    numeros = None
    try:
        elemento_imagem = wait.until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[1]/div[2]/div/div[2]/div[4]/div/div/div/span/img'))
        )
        driver.save_screenshot('screenshot.png')
        screenshot = Image.open('screenshot.png')
        location = elemento_imagem.location
        size = elemento_imagem.size
        left = location['x']
        top = location['y']
        right = left + size['width']
        bottom = top + size['height']
        imagem = screenshot.crop((left, top, right, bottom)).convert('L')
        numeros = pytesseract.image_to_string(imagem, config='--psm 6 -c tessedit_char_whitelist=0123456789')
        numeros = ''.join(filter(str.isdigit, numeros))
    except Exception as e:
        print(f"Erro na extração: {type(e).__name__}: {e}")
    return numeros


def preencher_campo(driver, element_id, valor, wait):
    try:
        campo = wait.until(EC.element_to_be_clickable((By.ID, element_id)))
        campo.clear()
        campo.send_keys(valor)
        print(f"Campo {element_id} preenchido")
    except Exception as e:
        print(f"Erro ao preencher {element_id}: {e}")


def processar_login(driver, wait):
    try:
        if driver.current_url == 'https://itapira.sigiss.com.br/itapira/contribuinte/main.php':
            return True
        current_url = driver.current_url
        if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
            print("Erro de login: Código de Confirmação Inválido (captcha)")
            return False
        if "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
            print("Erro de login: Contribuinte Inexistente ou Senha Inválida")
            return False
        try:
            erro_elemento = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[2]/div/font/b/center')))
            if erro_elemento.is_displayed():
                print("Erro no login: ", erro_elemento.text)
                return False
        except Exception:
            pass
    except Exception as e:
        print(f"Erro no processo de login: {e}")
    return False


def digitar_captcha(driver, numeros, wait):
    try:
        campo = wait.until(EC.element_to_be_clickable((By.ID, "confirma")))
        campo.clear()
        campo.send_keys(numeros)
        botao_logar = wait.until(EC.element_to_be_clickable((By.ID, "btnOk")))
        botao_logar.click()
        print("Captcha digitado com sucesso!")
    except Exception as e:
        print(f"Erro ao digitar captcha: {e}")


def atualizar_excel_status(linha_index, mensagem):
    try:
        workbook = load_workbook(CAMINHO_EXCEL)
        worksheet = workbook.active
        coluna_status = None
        for col_idx, col_name in enumerate(worksheet[1], 1):
            if col_name.value == 'Status Processo':
                coluna_status = col_idx
                break
        if coluna_status is None:
            coluna_status = worksheet.max_column + 1
            worksheet.cell(row=1, column=coluna_status, value='Status Processo')
        worksheet.cell(row=linha_index + 2, column=coluna_status, value=mensagem)
        workbook.save(CAMINHO_EXCEL)
        workbook.close()
    except Exception as e:
        print(f"Erro ao atualizar o Excel: {e}")


def nome_empresa_limpa(empresa):
    return re.sub(r'[<>:"/\\|?*]', '_', str(empresa or '').strip().upper())


def construir_pasta_notas_prestados(ano, mes, empresa):
    pasta = os.path.join(
        os.getcwd(),
        "notas_fiscais",
        "prestados",
        str(ano),
        str(mes).zfill(2),
        nome_empresa_limpa(empresa)
    )
    os.makedirs(pasta, exist_ok=True)
    return pasta


def aguardar_zip_download(pasta, timeout=120):
    fim = time.time() + timeout
    while time.time() < fim:
        arquivos = [
            f for f in glob.glob(os.path.join(pasta, "*.zip"))
            if not f.endswith(".crdownload")
        ]
        if arquivos:
            arquivos.sort(key=os.path.getmtime, reverse=True)
            return arquivos[0]
        time.sleep(1)
    return None


def extrair_zip(arquivo_zip, pasta_destino):
    try:
        with zipfile.ZipFile(arquivo_zip, 'r') as zip_ref:
            zip_ref.extractall(pasta_destino)
        os.remove(arquivo_zip)
        print(f"ZIP extraído em: {pasta_destino}")
        return True
    except zipfile.BadZipFile as e:
        print(f"ZIP inválido: {e}")
    except Exception as e:
        print(f"Não foi possível extrair o ZIP: {e}")
    return False


def exportar_notas_prestadas(driver, wait, mes, ano, empresa, linha_index=None):
    pasta_notas = construir_pasta_notas_prestados(ano, mes, empresa)
    try:
        click_element(wait, (By.XPATH, "//button[contains(text(), 'Acessórios')]"), "Botão 'Acessórios'")
        click_element(wait, (By.XPATH, "//a[contains(@onclick, \"abre_arquivo('dmm/_menu.php')\")]"), "Link 'Painel de Controle'")
        time.sleep(3)
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "main")))
        click_element(wait, (By.XPATH, "//td[contains(@onclick, \"display('nfe')\")]"), "Ferramentas NFS-e")
        click_element(wait, (By.XPATH, "//a[contains(@href, 'nfe_historico_exportacao.php')]"), "Link 'Exportar notas'")
        driver.switch_to.default_content()
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "main")))
        checkbox = wait.until(EC.element_to_be_clickable((By.ID, "todos")))
        if not checkbox.is_selected():
            checkbox.click()
        click_element(wait, (By.ID, "btnExportar"), "Botão 'Exportar xml'")
        alert = WebDriverWait(driver, 20).until(EC.alert_is_present())
        print("Alert de exportação encontrado.")
        alert.accept()
        arquivo_zip = aguardar_zip_download(pasta_notas, timeout=120)
        if not arquivo_zip:
            raise Exception("Nenhum arquivo ZIP encontrado após exportação.")
        extrair_zip(arquivo_zip, pasta_notas)
    except Exception as exc:
        print(f"Erro ao exportar notas: {exc}")
        if linha_index is not None:
            atualizar_excel_status(linha_index, 'Exportação de notas falhou')
    finally:
        driver.switch_to.default_content()


def preencher_data(driver, wait, mes, ano, empresa, linha_index=None):
    try:
        btn_alterar = wait.until(EC.element_to_be_clickable((By.ID, "btnAlterar")))
        btn_alterar.click()
        campo_mes = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[3]/select')))
        campo_mes.send_keys(mes)
        print(f"Mês '{mes}' digitado com sucesso!")
        campo_ano = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[7]/input')))
        campo_ano.clear()
        campo_ano.send_keys(ano)
        print(f"Ano '{ano}' digitado com sucesso!")
        botao_ok = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "btn-success")))
        botao_ok.click()
        print("Botão OK clicado com sucesso!")
        exportar_notas_prestadas(driver, wait, mes, ano, empresa, linha_index)
    except Exception as e:
        print(f"Erro ao preencher data: {e}")
        if linha_index is not None:
            atualizar_excel_status(linha_index, 'Não foi possível preencher data')


def main():
    try:
        df = pd.read_excel(CAMINHO_EXCEL, engine='openpyxl')
        for index, row in df.iterrows():
            tentativas = 0
            max_tentativas = 8
            login_bem_sucedido = False
            login_falhou_credenciais = False
            while tentativas < max_tentativas and not login_bem_sucedido:
                chrome_options = Options()
                pasta_notas = construir_pasta_notas_prestados(row['Ano'], row['Mês'], row['Empresa'])
                prefs = {
                    "download.default_directory": pasta_notas,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "plugins.plugins_disabled": ["Chrome PDF Viewer"],
                    "plugins.plugin_field_trial_triggered": False
                }
                chrome_options.add_experimental_option("prefs", prefs)
                driver = webdriver.Chrome(options=chrome_options)
                driver.maximize_window()
                wait = WebDriverWait(driver, 20)
                driver.get(URL_LOGIN)
                time.sleep(2)
                try:
                    btn_ciente = wait.until(EC.element_to_be_clickable((By.ID, "btnCiente")))
                    btn_ciente.click()
                    print("Botão 'Estou Ciente' clicado com sucesso!")
                except Exception:
                    print("Botão 'Estou Ciente' não encontrado ou já foi fechado.")
                try:
                    if tentativas == 0:
                        print(f"Processando linha {index + 1}: {row['Empresa']}")
                    else:
                        print(f"Tentativa {tentativas + 1} para a empresa: {row['Empresa']}")
                    preencher_campo(driver, "cnpj", row['Usuário'], wait)
                    preencher_campo(driver, "senha", row['Senha'], wait)
                    numeros = extrair_numeros_imagem(driver, wait)
                    if numeros:
                        print(f"Números extraídos: {numeros}")
                        digitar_captcha(driver, numeros, wait)
                        time.sleep(10)
                        if processar_login(driver, wait):
                            login_bem_sucedido = True
                            preencher_data(driver, wait, row['Mês'], row['Ano'], row['Empresa'], index)
                        else:
                            current_url = driver.current_url
                            if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
                                print(f"Captcha inválido para {row['Empresa']}, tentando novamente...")
                                tentativas += 1
                            elif "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
                                print(f"Login falhou por credenciais incorretas para {row['Empresa']}")
                                atualizar_excel_status(index, 'Não foi possivel realizar o login.')
                                login_falhou_credenciais = True
                                break
                            else:
                                print("Login falhou por outro motivo, não preenchendo a data.")
                                break
                    else:
                        print("Nenhum número foi detectado!")
                        tentativas += 1
                        if tentativas < max_tentativas:
                            print(f"Tentando novamente ({tentativas}/{max_tentativas})...")
                finally:
                    driver.quit()
            if not login_bem_sucedido and not login_falhou_credenciais:
                print(f"Excedido o número máximo de tentativas para {row['Empresa']}. Indo para a próxima empresa.")
    except Exception as e:
        print(f"Erro geral: {e}")


if __name__ == "__main__":
    main()
