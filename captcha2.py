from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import pytesseract
from PIL import Image
import pandas as pd
import time
import os
import glob
import re
from openpyxl import load_workbook

# Configurações
CAMINHO_TESSERACT = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
CAMINHO_EXCEL = r'C:\Users\bruno\Desktop\Automação\Resolve-Captcha\Senha Municipio Itapira Prestadoras (Maria).xlsx'
URL_LOGIN = 'https://itapira.sigiss.com.br/itapira/contribuinte/login.php'

pytesseract.pytesseract.tesseract_cmd = CAMINHO_TESSERACT

def extrair_numeros_imagem(driver, wait):
    numeros = None
    try:
        # Localizar o elemento da imagem
        elemento_imagem = wait.until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[1]/div[2]/div/div[2]/div[4]/div/div/div/span/img'))
        )

        # Tirar screenshot e processar
        driver.save_screenshot('screenshot.png')
        screenshot = Image.open('screenshot.png')
        
        # Calcular coordenadas
        location = elemento_imagem.location
        size = elemento_imagem.size
        left = location['x']
        top = location['y']
        right = location['x'] + size['width']
        bottom = location['y'] + size['height']

        # Recortar e processar imagem
        imagem = screenshot.crop((left, top, right, bottom))
        imagem.save('numero.png')
        
        # OCR com pré-processamento
        imagem = imagem.convert('L')  # Converter para escala de cinza
        numeros = pytesseract.image_to_string(imagem, config='--psm 6 -c tessedit_char_whitelist=0123456789')
        numeros = ''.join(filter(str.isdigit, numeros))

    except Exception as e:
        print(f"Erro na extração: {type(e).__name__}: {e}")
    
    return numeros

def preencher_campo(driver, element_id, valor, wait):
    try:
        campo = wait.until(
            EC.element_to_be_clickable((By.ID, element_id))
        )
        campo.clear()
        campo.send_keys(valor)
        print(f"Campo {element_id} preenchido")
    except Exception as e:
        print(f"Erro ao preencher {element_id}: {e}")

def processar_login(driver, wait):
    try:        
        # Verificar se o login foi bem-sucedido pelo URL
        if driver.current_url == 'https://itapira.sigiss.com.br/itapira/contribuinte/main.php':
            return True  # Login bem-sucedido

        # Verificar se há mensagem de erro de captcha inválido
        current_url = driver.current_url
        if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
            print("Erro de login: Código de Confirmação Inválido (captcha)")
            return False  # Indica que o login falhou devido a captcha incorreto

        # Verificar se há mensagem de erro de login específico (Contribuinte Inexistente ou Senha Inválida)
        if "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
            print("Erro de login: Contribuinte Inexistente ou Senha Inválida")
            return False  # Indica que o login falhou devido a credenciais incorretas

        # Verificar se há mensagem de erro genérica
        try:
            erro_elemento = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div[2]/div/font/b/center')))
            if erro_elemento.is_displayed():
                print("Erro no login: ", erro_elemento.text)
                # Verificar se o erro é relacionado ao captcha
                if "Confirmação" in erro_elemento.text or "confirmação" in erro_elemento.text or "Código" in erro_elemento.text:
                    return False  # Indica que o login falhou devido a captcha incorreto
                else:
                    return False  # Outro tipo de erro de login
        except Exception:
            pass  # Se não encontrar o elemento de erro, assume que não houve erro

    except Exception as e:
        print(f"Erro no processo de login: {e}")
        return False  # Indica que o login falhou


def digitar_captcha(driver, numeros, wait):
    try:
        # Localizar e interagir com o campo
        campo = wait.until(
            EC.element_to_be_clickable((By.ID, "confirma"))
        )
        campo.clear()
        campo.send_keys(numeros)
        print("Captcha digitado com sucesso!")
        
        botao_logar = wait.until(EC.element_to_be_clickable((By.ID, "btnOk")))
        botao_logar.click()
        
    except Exception as e:
        print(f"Erro ao digitar captcha: {e}")

def preencher_data(driver, wait, mes, ano, empresa):
    try:
        # Localizar e interagir com o campo
        campo_modificar = wait.until(
            EC.element_to_be_clickable((By.ID, "btnAlterar"))
        )
        campo_modificar.click()

        campo_mes = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[3]/select')))
        campo_mes.send_keys(mes)
        print(f"Mês '{mes}' digitado com sucesso!")
        
        campo_ano = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="panelFiltro"]/table/tbody/tr/td[7]/input')))
        campo_ano.clear()
        campo_ano.send_keys(ano)
        print(f"Ano '{ano}' digitado com sucesso!")
        
        # Clicar no botão OK após preencher mês e ano
        botao_ok = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "btn-success")))
        botao_ok.click()
        print("Botão OK clicado com sucesso!")
        
        # Após preencher data, verificar e clicar em Livro Fiscal para testar download (ou Encerramento Fiscal para produção)
        # Para testar o download, vamos usar a função de Livro Fiscal
        # clicar_livro_fiscal(driver, wait, mes, ano, empresa)
        # Para usar o encerramento fiscal (quando estiver disponível), substitua a linha acima por:
        clicar_encerramento_fiscal(driver, wait, mes, ano, empresa)
                
    except Exception as e:
        print(f"Erro ao preencher data: {e}")

def clicar_livro_fiscal(driver, wait, mes, ano, empresa):
    try:
        # Verificar se o botão "Serviços Prestados" existe
        servicos_prestados_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Serviços Prestados')]"))
        )
        
        # Clicar no botão "Serviços Prestados" para abrir o dropdown
        servicos_prestados_btn.click()
        print("Botão 'Serviços Prestados' clicado com sucesso!")
        
        # Aguardar o dropdown abrir e clicar em "Livro Fiscal"
        livro_fiscal_link = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@onclick, 'livroMensalP')]"))
        )
        livro_fiscal_link.click()
        print("Link 'Livro Fiscal' clicado com sucesso!")
        
        # Mudar para a nova aba que foi aberta com o PDF
        time.sleep(3)
        janelas = driver.window_handles
        if len(janelas) > 1:
            driver.switch_to.window(janelas[1])  # Mudar para a nova aba
            print("Mudou para a nova aba com o PDF do Livro Fiscal.")
            
            # Obter o caminho da pasta de download
            pasta_download = os.path.join(os.getcwd(), str(ano), str(mes).zfill(2))
            
            # Aguardar tempo suficiente para o download começar
            time.sleep(5)
            
            # Encontrar o arquivo PDF mais recente na pasta de download
            arquivos_pdf = glob.glob(os.path.join(pasta_download, "*.pdf"))
            if arquivos_pdf:
                # Ordenar arquivos por data de modificação (mais recente primeiro)
                arquivos_pdf.sort(key=os.path.getmtime, reverse=True)
                arquivo_original = arquivos_pdf[0]  # Pegar o mais recente
                
                # Criar novo nome para o arquivo usando o nome da empresa
                nome_limpo = re.sub(r'[<>:"/\\|?*]', '_', empresa)  # Remover caracteres inválidos para nomes de arquivo
                novo_nome = os.path.join(pasta_download, f"{nome_limpo}.pdf")
                
                # Renomear o arquivo
                os.rename(arquivo_original, novo_nome)
                print(f"Arquivo PDF renomeado para: {novo_nome}")
            else:
                print("Nenhum arquivo PDF encontrado na pasta de download.")
            
            # Fechar a aba do PDF
            driver.close()
            print("Fechou a aba do PDF do Livro Fiscal.")
            
            # Voltar para a aba principal
            driver.switch_to.window(janelas[0])
            print("Retornou para a aba principal.")
        else:
            print("Não foi possível abrir o PDF do Livro Fiscal em uma nova aba.")
        
    except Exception as e:
        print(f"Botão 'Serviços Prestados' não encontrado ou erro ao clicar em 'Livro Fiscal': {e}")

def clicar_encerramento_fiscal(driver, wait, mes, ano, empresa):
    try:
        # Verificar se o botão "Serviços Prestados" existe
        servicos_prestados_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Serviços Prestados')]"))
        )
        
        # Clicar no botão "Serviços Prestados" para abrir o dropdown
        servicos_prestados_btn.click()
        print("Botão 'Serviços Prestados' clicado com sucesso!")
        
        # Aguardar o dropdown abrir e clicar em "Encerramento Fiscal"
        encerramento_fiscal_link = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@onclick, 'fechamento/prestado.php')]"))
        )
        encerramento_fiscal_link.click()
        print("Link 'Encerramento Fiscal' clicado com sucesso!")
        
        # Aguardar o iframe carregar
        time.sleep(4)
        iframe = wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "main")))
        print("Entrou no iframe 'main'.")
        
        # Procurar e clicar no botão "Encerrar Mes"
        botao_encerrar = wait.until(
            EC.element_to_be_clickable((By.ID, "btnSalvar"))
        )
        botao_encerrar.click()
        print("Botão 'Encerrar Mes' clicado com sucesso!")
        
        # Lidar com o alerta do navegador com fallback de 2 minutos
        try:
            # Aguardar até 2 minutos pelo alerta
            from selenium.webdriver.support.ui import WebDriverWait
            fallback_wait = WebDriverWait(driver, 20)  # 2 minutos de espera
            alert = fallback_wait.until(EC.alert_is_present())
            alert_text = alert.text
            print(f"Alerta encontrado: {alert_text}")
            alert.accept()  # Clica em OK no alerta
            print("Alerta aceito com sucesso!")
        except:
            print("Nenhum alerta encontrado ou não foi possível interagir com ele após 2 minutos. Continuando para a próxima empresa.")
        
        # Após completar o encerramento fiscal, voltar ao frame principal
        driver.switch_to.default_content()
        print("Retornou ao frame principal.")
        
        # Aguardar um pouco para garantir que a página esteja estável
        time.sleep(2)
        
        # Clicar novamente no botão "Serviços Prestados"
        servicos_prestados_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Serviços Prestados')]"))
        )
        servicos_prestados_btn.click()
        print("Botão 'Serviços Prestados' clicado novamente com sucesso!")
        
        # Clicar no link "Histórico de Boletos"
        historico_boletos_link = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@onclick, 'relatorio/listaBoletos_prest.php')]"))
        )
        historico_boletos_link.click()
        print("Link 'Histórico de Boletos' clicado com sucesso!")
        
        # Aguardar o iframe carregar novamente
        time.sleep(4)
        iframe = wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "main")))
        print("Entrou no iframe 'main' para Histórico de Boletos.")
        
        # Preencher o select de mês com o valor do Excel
        select_mes = wait.until(
            EC.element_to_be_clickable((By.ID, "mesI"))
        )
        select_mes.send_keys(str(mes))
        print(f"Mês '{mes}' selecionado com sucesso!")
        
        # Preencher o input de ano com o valor do Excel
        input_ano = wait.until(
            EC.element_to_be_clickable((By.ID, "anoI"))
        )
        input_ano.clear()
        input_ano.send_keys(str(ano))
        print(f"Ano '{ano}' preenchido com sucesso!")
        
        # Clicar no botão "Filtrar"
        botao_filtrar = wait.until(
            EC.element_to_be_clickable((By.ID, "Submit"))
        )
        botao_filtrar.click()
        print("Botão 'Filtrar' clicado com sucesso!")
        
        # Aguardar o iframe "relatorio" carregar
        time.sleep(4)
        
        # Mudar para o iframe "relatorio"
        driver.switch_to.frame("relatorio")
        print("Entrou no iframe 'relatorio'.")
        
        # Procurar pelo link do boleto na segunda coluna (elemento com o número do boleto)
        try:
            # Localizar o link do boleto (na segunda coluna da tabela)
            link_boleto = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, 'javascript:boleto_ver')]"))
            )
            
            # Obter o href do link para construir a URL do PDF
            href = link_boleto.get_attribute("href")
            
            # Extrair os parâmetros do href (ex: boleto_ver(1481295, 9, 2025, 1, '', 1))
            # Vamos usar a lógica para construir a URL do PDF com base no padrão fornecido
            # O link real para o PDF é algo como:
            # https://itapira.sigiss.com.br/itapira/cgi-local/contribuinte/livro/livro_fiscal_mensal_prestado_pdf.php?ccm=6103&cnpj=14097460000194&mes=9&ano=2025
            
            # Como não temos os valores exatos de ccm e cnpj, vamos tentar obter o link real do PDF
            # Primeiro, vamos extrair o número do protocolo do boleto (20732083 no exemplo)
            numero_boleto = link_boleto.text
            print(f"Número do boleto encontrado: {numero_boleto}")
            
            # Em vez de tentar construir o URL do PDF, vamos tentar obter o link real do PDF
            # Primeiro, precisamos pegar o número do CCM da página ou da empresa logada
            # Por enquanto, faremos uma abordagem alternativa: clicar no link e esperar que o PDF seja baixado automaticamente
            # com as configurações do navegador
            
            # Clicar no link do boleto para abrir a nova aba com o PDF
            link_boleto.click()
            print("Link do boleto clicado com sucesso!")
            
            # Mudar para a nova aba que foi aberta
            time.sleep(3)
            janelas = driver.window_handles
            if len(janelas) > 1:
                driver.switch_to.window(janelas[1])  # Mudar para a nova aba
                print("Mudou para a nova aba com o PDF.")
                
                # Obter o caminho da pasta de download
                pasta_download = os.path.join(os.getcwd(), str(ano), str(mes).zfill(2))
                
                # Aguardar tempo suficiente para o download começar
                time.sleep(5)
                
                # Encontrar o arquivo PDF mais recente na pasta de download
                arquivos_pdf = glob.glob(os.path.join(pasta_download, "*.pdf"))
                if arquivos_pdf:
                    # Ordenar arquivos por data de modificação (mais recente primeiro)
                    arquivos_pdf.sort(key=os.path.getmtime, reverse=True)
                    arquivo_original = arquivos_pdf[0]  # Pegar o mais recente
                    
                    # Criar novo nome para o arquivo usando o nome da empresa
                    nome_limpo = re.sub(r'[<>:"/\\|?*]', '_', empresa)  # Remover caracteres inválidos para nomes de arquivo
                    novo_nome = os.path.join(pasta_download, f"{nome_limpo}.pdf")
                    
                    # Renomear o arquivo
                    os.rename(arquivo_original, novo_nome)
                    print(f"Arquivo PDF renomeado para: {novo_nome}")
                else:
                    print("Nenhum arquivo PDF encontrado na pasta de download.")
                
                # Fechar a aba do PDF
                driver.close()
                print("Fechou a aba do PDF.")
                
                # Voltar para a aba principal
                driver.switch_to.window(janelas[0])
                print("Retornou para a aba principal.")
            else:
                print("Não foi possível abrir o PDF em uma nova aba.")
            
        except:
            print("Nenhum boleto encontrado para o filtro selecionado. Continuando para a próxima empresa.")
        
        # Voltar ao frame principal
        driver.switch_to.default_content()
                
    except Exception as e:
        print(f"Botão 'Serviços Prestados' não encontrado ou erro ao clicar em 'Encerramento Fiscal': {e}")

def atualizar_excel_status(linha_index, mensagem):
    """Atualiza a coluna 'Status Processo' no Excel"""
    try:
        workbook = load_workbook(CAMINHO_EXCEL)
        worksheet = workbook.active
        
        # Assumindo que a coluna 'Status Processo' é a próxima após as existentes
        # Encontrar o índice da coluna 'Status Processo'
        coluna_status = None
        for col_idx, col_name in enumerate(worksheet[1], 1):
            if col_name.value == 'Status Processo':
                coluna_status = col_idx
                break
        
        if coluna_status is not None:
            # Atualizar a célula específica (linha_index + 2, pois o índice do Excel começa em 1 e a primeira linha é cabeçalho)
            worksheet.cell(row=linha_index + 2, column=coluna_status, value=mensagem)
        else:
            # Se a coluna não existir, adicionar como nova coluna
            coluna_status = worksheet.max_column + 1
            worksheet.cell(row=1, column=coluna_status, value='Status Processo')
            worksheet.cell(row=linha_index + 2, column=coluna_status, value=mensagem)
        
        workbook.save(CAMINHO_EXCEL)
        workbook.close()
    except Exception as e:
        print(f"Erro ao atualizar o Excel: {e}")

def main():
    try:
        # Ler dados do Excel
        df = pd.read_excel(CAMINHO_EXCEL, engine='openpyxl')

        # Para cada linha no Excel
        for index, row in df.iterrows():
            tentativas = 0
            max_tentativas = 8
            login_bem_sucedido = False
            login_falhou_credenciais = False  # Flag para indicar falha por credenciais
            
            while tentativas < max_tentativas and not login_bem_sucedido:
                # Configurar as opções do Chrome
                chrome_options = Options()
                
                # Criar pasta de download com estrutura ano/mes
                pasta_download = os.path.join(os.getcwd(), str(row['Ano']), str(row['Mês']).zfill(2))
                os.makedirs(pasta_download, exist_ok=True)
                
                # Configurar preferências de download
                prefs = {
                    "download.default_directory": pasta_download,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True  # Faz download em vez de abrir PDF no navegador
                }
                chrome_options.add_experimental_option("prefs", prefs)
                
                driver = webdriver.Chrome(options=chrome_options)
                driver.maximize_window()
                wait = WebDriverWait(driver, 20)
                driver.get(URL_LOGIN)
                
                time.sleep(2)  # Adiciona uma pausa para garantir que a página carregue completamente
                
                try:
                    if tentativas == 0:
                        print(f"Processando linha {index + 1}: {row['Empresa']}")
                    else:
                        print(f"Tentativa {tentativas + 1} para a empresa: {row['Empresa']}")
                    
                    # Preencher credenciais antes de extrair o captcha
                    preencher_campo(driver, "cnpj", row['Usuário'], wait)
                    preencher_campo(driver, "senha", row['Senha'], wait)
                    
                    # Extrair números
                    numeros = extrair_numeros_imagem(driver, wait)
                    
                    if numeros:
                        print(f"Números extraídos: {numeros}")
                        # Digitar no campo
                        digitar_captcha(driver, numeros, wait)
                        time.sleep(10)
                        
                        if processar_login(driver, wait):
                            login_bem_sucedido = True  # Define como True para sair do loop
                            preencher_data(driver, wait, row['Mês'], row['Ano'], row['Empresa'])  # Chama preencher_data apenas se o login for bem-sucedido
                        else:
                            # Verificar se o erro foi por captcha inválido
                            current_url = driver.current_url
                            if "msg=C%F3digo+de+Confirma%E7%E3o+Inv%E1lido" in current_url:
                                print(f"Captcha inválido para {row['Empresa']}, tentando novamente...")
                                tentativas += 1
                            elif "msg=Contribuinte+Inexistente+ou+Senha+Inv%E1lida" in current_url:
                                # Login falhou por credenciais incorretas
                                print(f"Login falhou por credenciais incorretas para {row['Empresa']}")
                                atualizar_excel_status(index, 'Não foi possivel realizar o login.')
                                login_falhou_credenciais = True
                                break  # Sai do loop e vai para a próxima empresa
                            else:
                                # Outro tipo de erro, não relacionado ao captcha
                                print("Login falhou por outro motivo, não preenchendo a data.")
                                break  # Sai do loop e vai para a próxima empresa
                    else:
                        print("Nenhum número foi detectado!")
                        tentativas += 1  # Tenta novamente, pois o captcha não foi extraído
                        if tentativas < max_tentativas:
                            print(f"Tentando novamente ({tentativas}/{max_tentativas})...")
                finally:
                    driver.quit()
            
            if not login_bem_sucedido and not login_falhou_credenciais:
                print(f"Excedido o número máximo de tentativas para {row['Empresa']}. Indo para a próxima empresa.")
                
    except Exception as e:
        print(f"Erro geral: {e}")

if __name__ == "__main__":
    main()

    